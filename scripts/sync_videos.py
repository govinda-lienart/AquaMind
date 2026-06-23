"""
Reads video metadata from an Excel sheet and registers it in MySQL.

Input  : XLSX_PATH spreadsheet with one video per row
Needs  : MySQL running, videos table created
Output : rows inserted into videos table (skips already-registered entries)
"""

# ── IMPORTS ───────────────────────────────────────────────────────────────────

import logging
from typing import Any

import openpyxl # reads excel files

from scripts.db import register_video

# ── CONSTANTS ─────────────────────────────────────────────────────────────────

logger = logging.getLogger(__name__)


# ── CONSTANTS ─────────────────────────────────────────────────────────────────

XLSX_PATH = 'video_metadata.xlsx'


# ── HELPERS ───────────────────────────────────────────────────────────────────

def load_video_rows(xlsx_path: str) -> list[dict[str, Any]]: # e.g. [{"file_path": "videos/IMG_0350.MOV", "fish_count": 5, ...},{"file_path": "videos/IMG_0651...]  # row 2 of xlsx

    wb      = openpyxl.load_workbook(xlsx_path) # the entire file
    ws      = wb.active # ws is worksheet - single sheet
    headers = [cell.value for cell in ws[1]] # select row 1 as tuple (all headers) # iterates over each of cell objects (headers)  eg # file_path | fish_count | -> 'file_path', 'fps', 'fish_count
    rows = ws.iter_rows(min_row=2, values_only=True) #generator object - getting the data rows start from row 2 - give aw value -> ('videos/IMG_0350.MOV', 5, 60, 'tracking', ...)
    return [dict(zip(headers, row)) for row in rows] #  pairs each header name with its matching value by position [  {'file_path': 'videos/IMG_0350.MOV', 'fish_count': 5},  # row 2...

# ── MAIN ──────────────────────────────────────────────────────────────────────

def main() -> None:

    rows    = load_video_rows(XLSX_PATH)
    added   = 0
    skipped = 0

    for data in rows:
        video_id = register_video(
            file_path      = data['file_path'],
            session_type   = data['session_type'],
            obstacles      = bool(data['obstacles']),
            fish_count     = data['fish_count'],
            notes          = data['notes'],
            species        = data['species'],
            morph          = data['morph'],
            tank_width_cm  = data['tank_width_cm'],
            tank_height_cm = data['tank_height_cm'],
            tank_depth_cm  = data['tank_depth_cm'],
            filmed_at      = data.get('filmed_at'),
        )

        if video_id:
            logger.info(f"added   {data['file_path']} → video_id={video_id}")
            added += 1
        else:
            logger.info(f"skipped {data['file_path']} (already registered)")
            skipped += 1

    logger.info(f"done — {added} added, {skipped} skipped")


# ── ENTRY POINT ───────────────────────────────────────────────────────────────

if __name__ == '__main__':
    from scripts.logger import setup_logging
    setup_logging()
    main()


