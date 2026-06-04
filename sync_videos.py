import openpyxl
from db import register_video

wb = openpyxl.load_workbook('video_metadata_2026_5_30.xlsx')
ws = wb.active

headers = [cell.value for cell in ws[1]]

added   = 0
skipped = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    data = dict(zip(headers, row))

    video_id = register_video(
        file_path      = data['file_path'],
        session_type   = data['session_type'],
        obstacles      = bool(data['obstacles']),
        fish_count     = data['fish_count'],
        notes          = data['notes'],
        species        = data['species'],
        morph          = data['morph'],
        tank_width_cm  = data['tank_width_cm'],
        tank_height_cm = data['tank_height_cm'],
        tank_depth_cm  = data['tank_depth_cm'],
        filmed_at      = data.get('filmed_at'),
    )

    if video_id:
        print(f"Added   {data['file_path']} → video_id={video_id}")
        added += 1
    else:
        print(f"Skipped {data['file_path']} (already registered)")
        skipped += 1

print(f"\nDone — {added} added, {skipped} skipped")
